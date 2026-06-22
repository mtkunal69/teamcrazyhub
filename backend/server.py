from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import uuid
import jwt
import bcrypt
import random
import logging
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal

from fastapi import FastAPI, APIRouter, Depends, HTTPException, Header, Query, UploadFile, File, Form, Request
from fastapi.responses import StreamingResponse
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, ConfigDict
from apscheduler.schedulers.asyncio import AsyncIOScheduler
from apscheduler.triggers.cron import CronTrigger
import pytz
import io
import tempfile
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from telegram_service import (
    send_message as tg_send,
    build_report_notification,
    build_daily_summary,
    get_config as tg_get_config,
    set_webhook as tg_set_webhook,
    delete_webhook as tg_delete_webhook,
    get_webhook_info as tg_get_webhook_info,
    get_me as tg_get_me,
    get_chat as tg_get_chat,
    create_invite_link as tg_create_invite,
    revoke_invite_link as tg_revoke_invite,
    approve_join_request as tg_approve_join,
)
import httpx
from gemini_service import (
    analyze_video as gemini_analyze,
    test_api_key as gemini_test_key,
    get_gemini_config,
)

# ─────────────────────────────────────────────────────────────
# Config & DB
# ─────────────────────────────────────────────────────────────
MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]
JWT_SECRET = os.environ["JWT_SECRET"]
JWT_ALGO = "HS256"
JWT_EXPIRES_MIN = 60 * 24 * 7  # 7 days

ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "admin123")
ADMIN_NAME = os.environ.get("ADMIN_NAME", "TeamCrazy Admin")

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI(title="TeamCrazy Hub API")
api = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format="%(asctime)s | %(levelname)s | %(message)s")
log = logging.getLogger("tch")

# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────
def hash_pw(p: str) -> str:
    return bcrypt.hashpw(p.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")

def verify_pw(p: str, h: str) -> bool:
    return bcrypt.checkpw(p.encode("utf-8"), h.encode("utf-8"))

def make_token(subject: str, role: str, name: str) -> str:
    payload = {
        "sub": subject,
        "role": role,
        "name": name,
        "exp": datetime.now(timezone.utc) + timedelta(minutes=JWT_EXPIRES_MIN),
        "iat": datetime.now(timezone.utc),
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGO)

def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGO])
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")

def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()

async def get_principal(authorization: Optional[str] = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Not authenticated")
    return decode_token(authorization[7:])

async def require_admin(p: dict = Depends(get_principal)) -> dict:
    if p.get("role") != "admin":
        raise HTTPException(403, "Admin only")
    return p

async def require_user(p: dict = Depends(get_principal)) -> dict:
    if p.get("role") != "user":
        raise HTTPException(403, "User only")
    return p

def strip_id(doc: dict) -> dict:
    if doc is None:
        return doc
    doc.pop("_id", None)
    return doc

# ─────────────────────────────────────────────────────────────
# Models
# ─────────────────────────────────────────────────────────────
WORKER_TYPES = ["4 ID Worker", "5 ID Worker", "7 ID Worker", "10 ID Worker"]

class AdminLogin(BaseModel):
    username: str
    password: str

class UserLogin(BaseModel):
    username: str
    password: str

class UserSignup(BaseModel):
    username: str  # platform username (login)
    password: str
    name: str  # full name
    telegram: str  # @handle
    default_worker_type: str = "5 ID Worker"

class ChannelIn(BaseModel):
    chat_id: Optional[str] = None  # numeric ID or @username
    invite_link: Optional[str] = None  # e.g. https://t.me/+abc or https://t.me/channelname
    title_override: Optional[str] = None  # admin can override display name
    default_worker_type: Optional[str] = None
    notes: Optional[str] = None

class ChannelPatch(BaseModel):
    title_override: Optional[str] = None
    default_worker_type: Optional[str] = None
    notes: Optional[str] = None
    enabled: Optional[bool] = None

class SlabIn(BaseModel):
    worker_type: str
    min: int
    max: Optional[int] = None  # None == unlimited
    salary: int
    enabled: bool = True

class SlabPatch(BaseModel):
    min: Optional[int] = None
    max: Optional[int] = None  # use -1 to signal unlimited via PUT? keep None=unchanged; use "unlimited" flag
    unlimited: Optional[bool] = None
    salary: Optional[int] = None
    enabled: Optional[bool] = None

class ReportIn(BaseModel):
    worker_type: str
    member_count: int
    video_filename: Optional[str] = None

class AICountIn(BaseModel):
    entered_count: int
    worker_type: str

class TelegramSettings(BaseModel):
    bot_token: str = ""
    chat_id: str = ""
    enabled: bool = False
    notify_on_report: bool = True
    daily_summary: bool = True

class GeminiSettings(BaseModel):
    api_key: str = ""
    model: str = "gemini-2.5-flash"
    enabled: bool = False

class LinkCreateIn(BaseModel):
    channel: str  # @username OR numeric chat_id
    name: Optional[str] = None  # optional link label

class ReportIn(BaseModel):
    worker_type: str
    link_id: Optional[str] = None  # invite link to count
    member_count: Optional[int] = None  # manual fallback if no link

# ─────────────────────────────────────────────────────────────
# Seed defaults
# ─────────────────────────────────────────────────────────────
DEFAULT_SLABS = {
    "4 ID Worker": [(0,39,0),(40,49,150),(50,59,220),(60,74,280),(75,None,350)],
    "5 ID Worker": [(0,49,0),(50,64,200),(65,79,300),(80,99,400),(100,None,500)],
    "7 ID Worker": [(0,59,0),(60,79,250),(80,99,380),(100,124,500),(125,None,650)],
    "10 ID Worker":[(0,79,0),(80,99,300),(100,124,450),(125,149,600),(150,None,800)],
}

def slab_label(mn: int, mx: Optional[int]) -> str:
    return f"{mn}+" if mx is None else f"{mn}\u2013{mx}"

async def seed_admin():
    existing = await db.admins.find_one({"username": ADMIN_USERNAME})
    if not existing:
        await db.admins.insert_one({
            "id": str(uuid.uuid4()),
            "username": ADMIN_USERNAME,
            "password_hash": hash_pw(ADMIN_PASSWORD),
            "name": ADMIN_NAME,
            "created_at": now_iso(),
        })
        log.info(f"Seeded admin: {ADMIN_USERNAME}")
    else:
        # keep password in sync with env
        if not verify_pw(ADMIN_PASSWORD, existing["password_hash"]):
            await db.admins.update_one(
                {"username": ADMIN_USERNAME},
                {"$set": {"password_hash": hash_pw(ADMIN_PASSWORD)}},
            )
            log.info("Updated admin password from env")

async def seed_slabs():
    count = await db.salary_slabs.count_documents({})
    if count > 0:
        return
    docs = []
    for wt, rows in DEFAULT_SLABS.items():
        for order, (mn, mx, sal) in enumerate(rows):
            docs.append({
                "id": str(uuid.uuid4()),
                "worker_type": wt,
                "min": mn,
                "max": mx,
                "salary": sal,
                "label": slab_label(mn, mx),
                "enabled": True,
                "order": order,
                "created_at": now_iso(),
                "updated_at": now_iso(),
            })
    await db.salary_slabs.insert_many(docs)
    log.info(f"Seeded {len(docs)} default salary slabs")

async def ensure_indexes():
    await db.admins.create_index("username", unique=True)
    # Drop legacy indexes on users (telegram unique → migrate to sparse + add username unique)
    try:
        existing = await db.users.index_information()
        if "telegram_1" in existing and existing["telegram_1"].get("unique"):
            await db.users.drop_index("telegram_1")
    except Exception:
        pass
    await db.users.create_index("username", unique=True, sparse=True)
    await db.users.create_index("telegram", sparse=True)
    await db.salary_slabs.create_index([("worker_type", 1), ("order", 1)])
    await db.reports.create_index([("created_at", -1)])
    await db.reports.create_index("telegram")
    await db.audit_logs.create_index([("created_at", -1)])
    await db.channels.create_index("chat_id", unique=True, sparse=True)
    await db.invite_links.create_index("link_url", unique=True, sparse=True)
    await db.link_members.create_index([("link_id", 1), ("telegram_user_id", 1)], unique=True)
    await db.link_members.create_index([("channel_id", 1), ("telegram_user_id", 1)])

# ─────────────────────────────────────────────────────────────
# Salary calc
# ─────────────────────────────────────────────────────────────
async def calc_salary(worker_type: str, count: int) -> tuple[int, Optional[str]]:
    slabs = await db.salary_slabs.find(
        {"worker_type": worker_type, "enabled": True}, {"_id": 0}
    ).sort("min", 1).to_list(100)
    for s in slabs:
        mx = s["max"] if s["max"] is not None else 10**9
        if s["min"] <= count <= mx:
            return s["salary"], s["label"]
    return 0, None

# ─────────────────────────────────────────────────────────────
# Routes — Auth
# ─────────────────────────────────────────────────────────────
@api.post("/auth/admin/login")
async def admin_login(body: AdminLogin):
    a = await db.admins.find_one({"username": body.username})
    if not a or not verify_pw(body.password, a["password_hash"]):
        raise HTTPException(401, "Invalid credentials")
    token = make_token(a["id"], "admin", a["name"])
    return {"token": token, "role": "admin", "name": a["name"], "username": a["username"]}

@api.post("/auth/user/register")
async def user_register(body: UserSignup):
    uname = body.username.strip().lower()
    if len(uname) < 3:
        raise HTTPException(400, "Username must be at least 3 characters")
    if len(body.password) < 6:
        raise HTTPException(400, "Password must be at least 6 characters")
    if await db.users.find_one({"username": uname}):
        raise HTTPException(409, "Username already taken")
    tg = body.telegram.strip()
    if not tg.startswith("@"):
        tg = "@" + tg
    if body.default_worker_type not in WORKER_TYPES:
        raise HTTPException(400, "Invalid worker type")
    user = {
        "id": str(uuid.uuid4()),
        "username": uname,
        "password_hash": hash_pw(body.password),
        "name": body.name.strip(),
        "telegram": tg,
        "default_worker_type": body.default_worker_type,
        "created_at": now_iso(),
        "last_login": now_iso(),
    }
    await db.users.insert_one(user)
    token = make_token(user["id"], "user", user["name"])
    return {"token": token, "role": "user", "name": user["name"], "username": uname, "telegram": tg, "user_id": user["id"]}

@api.post("/auth/user/login")
async def user_login(body: UserLogin):
    uname = body.username.strip().lower()
    u = await db.users.find_one({"username": uname})
    if not u or not u.get("password_hash") or not verify_pw(body.password, u["password_hash"]):
        raise HTTPException(401, "Invalid username or password")
    await db.users.update_one({"id": u["id"]}, {"$set": {"last_login": now_iso()}})
    token = make_token(u["id"], "user", u["name"])
    return {"token": token, "role": "user", "name": u["name"], "username": uname, "telegram": u["telegram"], "user_id": u["id"]}

@api.get("/auth/me")
async def me(p: dict = Depends(get_principal)):
    return {"id": p["sub"], "role": p["role"], "name": p["name"]}

# ─────────────────────────────────────────────────────────────
# Routes — Salary Charts
# ─────────────────────────────────────────────────────────────
@api.get("/charts")
async def get_charts():
    slabs = await db.salary_slabs.find({}, {"_id": 0}).sort([("worker_type",1),("min",1)]).to_list(1000)
    out = {wt: [] for wt in WORKER_TYPES}
    for s in slabs:
        if s["worker_type"] in out:
            out[s["worker_type"]].append(s)
    return out

async def push_audit(admin: dict, worker_type: str, action: str, slab_label_: str,
                    old_value: str = "", new_value: str = ""):
    await db.audit_logs.insert_one({
        "id": str(uuid.uuid4()),
        "admin_id": admin["sub"],
        "admin_name": admin["name"],
        "worker_type": worker_type,
        "action": action,
        "slab_label": slab_label_,
        "old_value": old_value,
        "new_value": new_value,
        "created_at": now_iso(),
    })

@api.post("/charts/slab")
async def add_slab(body: SlabIn, admin: dict = Depends(require_admin)):
    if body.worker_type not in WORKER_TYPES:
        raise HTTPException(400, "Invalid worker type")
    label = slab_label(body.min, body.max)
    doc = {
        "id": str(uuid.uuid4()),
        "worker_type": body.worker_type,
        "min": body.min, "max": body.max, "salary": body.salary,
        "label": label, "enabled": body.enabled, "order": 999,
        "created_at": now_iso(), "updated_at": now_iso(),
    }
    await db.salary_slabs.insert_one(doc)
    await push_audit(admin, body.worker_type, "ADD", label, "", f"\u20b9{body.salary}")
    doc.pop("_id", None)
    return doc

@api.put("/charts/slab/{slab_id}")
async def edit_slab(slab_id: str, body: SlabPatch, admin: dict = Depends(require_admin)):
    old = await db.salary_slabs.find_one({"id": slab_id}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Slab not found")
    upd = {}
    new_min = old["min"] if body.min is None else body.min
    new_max = old["max"]
    if body.unlimited is True:
        new_max = None
    elif body.max is not None:
        new_max = body.max
    new_sal = old["salary"] if body.salary is None else body.salary
    upd["min"] = new_min
    upd["max"] = new_max
    upd["salary"] = new_sal
    upd["label"] = slab_label(new_min, new_max)
    if body.enabled is not None:
        upd["enabled"] = body.enabled
    upd["updated_at"] = now_iso()
    await db.salary_slabs.update_one({"id": slab_id}, {"$set": upd})
    await push_audit(admin, old["worker_type"], "EDIT", upd["label"],
                     f"Range:{old['label']} | \u20b9{old['salary']}",
                     f"Range:{upd['label']} | \u20b9{new_sal}")
    new = await db.salary_slabs.find_one({"id": slab_id}, {"_id": 0})
    return new

@api.patch("/charts/slab/{slab_id}/toggle")
async def toggle_slab(slab_id: str, admin: dict = Depends(require_admin)):
    old = await db.salary_slabs.find_one({"id": slab_id}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Slab not found")
    new_state = not old["enabled"]
    await db.salary_slabs.update_one({"id": slab_id}, {"$set": {"enabled": new_state, "updated_at": now_iso()}})
    await push_audit(admin, old["worker_type"], "DISABLE" if not new_state else "ENABLE",
                     old["label"], "", "Disabled" if not new_state else "Enabled")
    return {"id": slab_id, "enabled": new_state}

@api.delete("/charts/slab/{slab_id}")
async def del_slab(slab_id: str, admin: dict = Depends(require_admin)):
    old = await db.salary_slabs.find_one({"id": slab_id}, {"_id": 0})
    if not old:
        raise HTTPException(404, "Slab not found")
    await db.salary_slabs.delete_one({"id": slab_id})
    await push_audit(admin, old["worker_type"], "DELETE", old["label"], f"\u20b9{old['salary']}", "")
    return {"ok": True}

@api.post("/charts/reset")
async def reset_charts(admin: dict = Depends(require_admin)):
    await db.salary_slabs.delete_many({})
    await seed_slabs()
    await push_audit(admin, "ALL", "RESET", "All", "", "Reset to defaults")
    return {"ok": True}

# ─────────────────────────────────────────────────────────────
# Routes — AI Count + Reports
# ─────────────────────────────────────────────────────────────
SAMPLE_NAMES = [
    "Anurag Dwivedi","BLACK KEY","MR. S.P. SIR","ISHIKA","Rahul Sharma","Priya Singh",
    "Amit Kumar","Sneha Patel","Vikram Rao","Neha Gupta","Arjun Mehta","Pooja Verma",
    "Suresh Nair","Kavita Joshi","Rohit Das","Anita Roy","Deepak Tiwari","Meena Shah",
    "Ravi Pillai","Sunita Yadav","Manish Dubey","Rekha Mishra","Ajay Sinha","Geeta Bose",
]

@api.post("/ai/count")
async def ai_count(body: AICountIn, _: dict = Depends(require_user)):
    """Simulated AI vision: scans 'video' and produces its own member count.
    Behavior:
      - 25% chance: AI detects significantly different count → MISMATCH
      - 75% chance: AI count within ±2 of entered → VERIFIED
    """
    entered = body.entered_count
    roll = random.random()
    if roll < 0.25:
        # Mismatch case — AI sees noticeably different count
        delta_pct = random.uniform(0.15, 0.40)
        sign = 1 if random.random() > 0.5 else -1
        ai_n = max(0, int(entered * (1 + sign * delta_pct)))
        if ai_n == entered:  # ensure actual mismatch
            ai_n = max(0, entered + (5 * sign))
    else:
        # Verified case — small variance
        ai_n = max(0, entered + random.randint(-2, 2))

    # Detect "premium" members for display
    total_visible = ai_n
    premium_count = max(0, int(total_visible * (0.15 + random.random() * 0.15)))
    members = []
    if total_visible > 0:
        premium_idx = set(random.sample(range(total_visible), min(premium_count, total_visible)))
        for i in range(total_visible):
            is_premium = i in premium_idx
            members.append({
                "name": random.choice(SAMPLE_NAMES),
                "premium": is_premium,
                "badge": ("\u2b50 Premium" if random.random() > 0.5 else "\U0001f381 Gift Badge") if is_premium else None,
            })

    matched = abs(ai_n - entered) <= 3
    return {
        "total_visible": total_visible,
        "premium_count": premium_count,
        "ai_count": ai_n,
        "matched": matched,
        "members": members[:15],
    }

@api.post("/reports")
async def submit_report(body: ReportIn, p: dict = Depends(require_user)):
    """Submit a daily salary report.
    - If link_id provided: count comes from invite_links.members_joined (real-time Telegram data)
    - Else: uses provided member_count (manual fallback, marked as MANUAL source)
    """
    if body.worker_type not in WORKER_TYPES:
        raise HTTPException(400, "Invalid worker type")

    user = await db.users.find_one({"id": p["sub"]}, {"_id": 0})
    if not user:
        raise HTTPException(404, "User not found")

    channel_title = None
    source = "MANUAL"
    if body.link_id:
        link = await db.invite_links.find_one({"id": body.link_id, "staff_user_id": p["sub"]}, {"_id": 0})
        if not link:
            raise HTTPException(404, "Link not found or not yours")
        ai_count_val = int(link.get("members_joined", 0))
        channel_title = link.get("channel_title")
        source = "INVITE_LINK"
    elif body.member_count is not None:
        ai_count_val = int(body.member_count)
    else:
        raise HTTPException(400, "Provide either link_id or member_count")

    # Status: with invite link, always VERIFIED (count is authoritative from Telegram)
    # With manual: always VERIFIED too (no separate AI to compare with)
    status = "VERIFIED"
    salary, slab = await calc_salary(body.worker_type, ai_count_val)

    doc = {
        "id": str(uuid.uuid4()),
        "user_id": p["sub"],
        "name": user["name"],
        "telegram": user["telegram"],
        "worker_type": body.worker_type,
        "member_count": ai_count_val,
        "ai_count": ai_count_val,
        "source": source,
        "link_id": body.link_id,
        "channel_title": channel_title,
        "salary": salary,
        "status": status,
        "slab_label": slab,
        "created_at": now_iso(),
    }
    await db.reports.insert_one(doc)
    doc.pop("_id", None)

    try:
        cfg = await tg_get_config(db)
        if cfg and cfg.get("enabled") and cfg.get("notify_on_report", True):
            await tg_send(db, build_report_notification(doc))
    except Exception as e:
        log.warning(f"TG notify failed: {e}")
    return doc

@api.get("/reports")
async def list_reports(
    admin: dict = Depends(require_admin),
    worker_type: Optional[str] = None,
    status: Optional[str] = None,
    search: Optional[str] = None,
    date: Optional[str] = None,  # YYYY-MM-DD
):
    q = {}
    if worker_type and worker_type != "ALL":
        q["worker_type"] = worker_type
    if status and status != "ALL":
        q["status"] = status
    if search:
        q["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"telegram": {"$regex": search, "$options": "i"}},
        ]
    if date:
        q["created_at"] = {"$gte": date, "$lt": date + "T99:99:99"}
    rows = await db.reports.find(q, {"_id": 0}).sort("created_at", -1).to_list(2000)
    return rows

@api.get("/reports/me")
async def my_reports(p: dict = Depends(require_user)):
    rows = await db.reports.find({"user_id": p["sub"]}, {"_id": 0}).sort("created_at", -1).to_list(500)
    return rows

@api.get("/reports/{rid}")
async def get_report(rid: str, p: dict = Depends(get_principal)):
    r = await db.reports.find_one({"id": rid}, {"_id": 0})
    if not r:
        raise HTTPException(404, "Not found")
    if p["role"] == "user" and r["user_id"] != p["sub"]:
        raise HTTPException(403, "Forbidden")
    return r

# ─────────────────────────────────────────────────────────────
# Routes — Audit + Dashboard
# ─────────────────────────────────────────────────────────────
@api.get("/audit")
async def audit(admin: dict = Depends(require_admin), action: Optional[str] = None):
    q = {} if not action or action == "ALL" else {"action": action}
    rows = await db.audit_logs.find(q, {"_id": 0}).sort("created_at", -1).to_list(1000)
    return rows

@api.get("/dashboard/stats")
async def dashboard_stats(admin: dict = Depends(require_admin)):
    reports = await db.reports.find({}, {"_id": 0}).to_list(5000)
    today = datetime.now(timezone.utc).date().isoformat()
    today_count = sum(1 for r in reports if r["created_at"][:10] == today)
    total_paid = sum(r["salary"] for r in reports)
    verified = sum(1 for r in reports if r["status"] == "VERIFIED")
    mismatch = sum(1 for r in reports if r["status"] == "MISMATCH COUNTING")
    by_type = []
    for wt in WORKER_TYPES:
        type_rows = [r for r in reports if r["worker_type"] == wt]
        by_type.append({
            "worker_type": wt,
            "count": len(type_rows),
            "paid": sum(r["salary"] for r in type_rows),
        })
    active_slabs = await db.salary_slabs.count_documents({"enabled": True})
    return {
        "total_reports": len(reports),
        "today_reports": today_count,
        "total_paid": total_paid,
        "verified": verified,
        "mismatch": mismatch,
        "active_slabs": active_slabs,
        "by_type": by_type,
        "recent": sorted(reports, key=lambda r: r["created_at"], reverse=True)[:5],
    }

@api.get("/health")
async def health():
    return {"status": "ok", "time": now_iso()}

# ─────────────────────────────────────────────────────────────
# Routes — Excel Export
# ─────────────────────────────────────────────────────────────
def _parse_iso(s: str) -> datetime:
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        return datetime.now(timezone.utc)

def _build_xlsx(reports: list, period_label: str, start_dt: datetime, end_dt: datetime) -> bytes:
    wb = Workbook()
    HDR_FILL = PatternFill("solid", fgColor="1D4ED8")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUB_FILL = PatternFill("solid", fgColor="EFF6FF")
    SUB_FONT = Font(bold=True, color="1D4ED8", size=11)
    THIN = Side(border_style="thin", color="E2E8F0")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER = Alignment(horizontal="center", vertical="center")

    # ── Sheet 1: Summary ──
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "TeamCrazy Hub · Weekly Salary Report"
    ws["A1"].font = Font(bold=True, size=16, color="1D4ED8")
    ws.merge_cells("A1:D1")
    ws["A2"] = f"Period: {period_label}"
    ws["A2"].font = Font(italic=True, color="64748B", size=10)
    ws.merge_cells("A2:D2")

    total_paid = sum(r.get("salary", 0) for r in reports)
    verified = sum(1 for r in reports if r.get("status") == "VERIFIED")
    mismatch = len(reports) - verified
    unique_workers = len({r.get("telegram") for r in reports})

    stats = [
        ("Total Reports", len(reports)),
        ("Unique Workers", unique_workers),
        ("Verified", verified),
        ("Mismatches", mismatch),
        ("Total Salary Paid", f"\u20b9{total_paid:,}"),
        ("Success Rate", f"{round(verified/max(len(reports),1)*100,1)}%"),
    ]
    ws["A4"] = "Metric"
    ws["B4"] = "Value"
    for col in ("A4", "B4"):
        ws[col].fill = HDR_FILL
        ws[col].font = HDR_FONT
        ws[col].border = BORDER
    for i, (k, v) in enumerate(stats, 5):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"A{i}"].border = BORDER
        ws[f"B{i}"].border = BORDER

    # By worker type
    row = 12
    ws[f"A{row}"] = "By Worker Type"
    ws[f"A{row}"].font = SUB_FONT
    ws[f"A{row}"].fill = SUB_FILL
    ws.merge_cells(f"A{row}:D{row}")
    row += 1
    ws[f"A{row}"] = "Worker Type"
    ws[f"B{row}"] = "Reports"
    ws[f"C{row}"] = "Total Salary"
    for c in ("A", "B", "C"):
        cell = ws[f"{c}{row}"]
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER
        cell.alignment = CENTER
    row += 1
    for wt in WORKER_TYPES:
        rows_wt = [r for r in reports if r.get("worker_type") == wt]
        ws[f"A{row}"] = wt
        ws[f"B{row}"] = len(rows_wt)
        ws[f"C{row}"] = f"\u20b9{sum(r.get('salary',0) for r in rows_wt):,}"
        for c in ("A", "B", "C"):
            ws[f"{c}{row}"].border = BORDER
        row += 1

    # ── Sheet 2: All Reports ──
    ws2 = wb.create_sheet("All Reports")
    headers = ["Date", "Name", "Telegram", "Worker Type", "Entered Count", "AI Count", "Slab", "Salary (\u20b9)", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER
        cell.alignment = CENTER

    sorted_reports = sorted(reports, key=lambda x: x.get("created_at", ""), reverse=True)
    for ri, r in enumerate(sorted_reports, 2):
        dt = _parse_iso(r.get("created_at", "")) + timedelta(hours=5, minutes=30)
        vals = [
            dt.strftime("%d %b %Y · %I:%M %p IST"),
            r.get("name", ""),
            r.get("telegram", ""),
            r.get("worker_type", ""),
            r.get("member_count", 0),
            r.get("ai_count", 0),
            r.get("slab_label") or "\u2014",
            r.get("salary", 0),
            r.get("status", ""),
        ]
        for ci, v in enumerate(vals, 1):
            cell = ws2.cell(row=ri, column=ci, value=v)
            cell.border = BORDER
            if ci == 9:
                cell.font = Font(bold=True, color=("047857" if v == "VERIFIED" else "B91C1C"))
            if ci == 8:
                cell.font = Font(bold=True, color="047857" if v > 0 else "94A3B8")

    # ── Sheet 3: Per-Worker Breakdown ──
    ws3 = wb.create_sheet("Worker Breakdown")
    for col, h in enumerate(["Worker", "Telegram", "Reports", "Total Members", "Total Salary (\u20b9)", "Verified", "Mismatches"], 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.border = BORDER
        cell.alignment = CENTER

    by_user = {}
    for r in reports:
        k = r.get("telegram") or r.get("name") or "?"
        d = by_user.setdefault(k, {"name": r.get("name", ""), "tg": r.get("telegram", ""),
                                    "n": 0, "members": 0, "salary": 0, "v": 0, "m": 0})
        d["n"] += 1
        d["members"] += r.get("member_count", 0)
        d["salary"] += r.get("salary", 0)
        if r.get("status") == "VERIFIED":
            d["v"] += 1
        else:
            d["m"] += 1

    for ri, (_k, u) in enumerate(sorted(by_user.items(), key=lambda x: x[1]["salary"], reverse=True), 2):
        vals = [u["name"], u["tg"], u["n"], u["members"], u["salary"], u["v"], u["m"]]
        for ci, v in enumerate(vals, 1):
            cell = ws3.cell(row=ri, column=ci, value=v)
            cell.border = BORDER
            if ci == 5:
                cell.font = Font(bold=True, color="047857" if v > 0 else "94A3B8")

    # Auto-size columns
    for sheet in [ws, ws2, ws3]:
        for col_cells in sheet.columns:
            try:
                col_letter = col_cells[0].column_letter
                max_len = max(len(str(c.value or "")) for c in col_cells)
                sheet.column_dimensions[col_letter].width = min(max_len + 4, 38)
            except Exception:
                pass

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@api.get("/reports/export/weekly")
async def export_weekly_xlsx(admin: dict = Depends(require_admin), days: int = 7):
    end_dt = datetime.now(timezone.utc)
    start_dt = end_dt - timedelta(days=days)
    rows = await db.reports.find(
        {"created_at": {"$gte": start_dt.isoformat()}},
        {"_id": 0},
    ).sort("created_at", -1).to_list(10000)

    ist_now = datetime.now(IST)
    ist_start = ist_now - timedelta(days=days)
    period_label = f"{ist_start.strftime('%d %b %Y')} \u2192 {ist_now.strftime('%d %b %Y')}"
    xlsx_bytes = _build_xlsx(rows, period_label, start_dt, end_dt)

    filename = f"teamcrazy_weekly_report_{ist_now.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

# ─────────────────────────────────────────────────────────────
# Routes — Telegram Settings
# ─────────────────────────────────────────────────────────────
@api.get("/settings/telegram")
async def get_tg_settings(admin: dict = Depends(require_admin)):
    cfg = await db.settings.find_one({"key": "telegram"}, {"_id": 0})
    if not cfg:
        return {"bot_token": "", "chat_id": "", "enabled": False,
                "notify_on_report": True, "daily_summary": True}
    cfg.pop("key", None)
    return cfg

@api.put("/settings/telegram")
async def set_tg_settings(body: TelegramSettings, admin: dict = Depends(require_admin)):
    doc = body.model_dump()
    doc["key"] = "telegram"
    doc["updated_at"] = now_iso()
    doc["updated_by"] = admin["name"]
    await db.settings.update_one({"key": "telegram"}, {"$set": doc}, upsert=True)
    return {"ok": True, "saved": True}

@api.post("/settings/telegram/test")
async def test_tg(admin: dict = Depends(require_admin)):
    """Test message — bypasses the 'enabled' flag so admin can verify creds before going live."""
    cfg = await db.settings.find_one({"key": "telegram"}, {"_id": 0})
    if not cfg:
        return {"ok": False, "error": "Save bot token + chat ID first"}
    token = (cfg.get("bot_token") or "").strip()
    chat_id = (cfg.get("chat_id") or "").strip()
    if not token:
        return {"ok": False, "error": "Bot token is missing"}
    if not chat_id:
        return {"ok": False, "error": "Chat ID is missing"}
    text = (
        "\u2705 <b>TeamCrazy Hub</b> \u2014 test message\n"
        "Your Telegram bot is connected and working.\n"
        f"Sent by: <b>{admin['name']}</b> \u00b7 {now_iso()[:19]} UTC"
    )
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    try:
        async with httpx.AsyncClient(timeout=20) as client_tg:
            r = await client_tg.post(url, json={
                "chat_id": chat_id,
                "text": text,
                "parse_mode": "HTML",
                "disable_web_page_preview": True,
            })
            data = r.json()
            if not data.get("ok"):
                desc = data.get("description", "Unknown error from Telegram")
                return {"ok": False, "error": f"Telegram API: {desc}"}
            return {"ok": True, "message_id": data.get("result", {}).get("message_id")}
    except httpx.TimeoutException:
        return {"ok": False, "error": "Telegram API timeout — check VPS internet"}
    except Exception as e:
        return {"ok": False, "error": f"Network error: {str(e)[:120]}"}

@api.post("/settings/telegram/send-daily-now")
async def send_daily_now(admin: dict = Depends(require_admin)):
    """Manually trigger today's summary (useful for testing)."""
    await run_daily_summary_job(force=True)
    return {"ok": True}

# ─────────────────────────────────────────────────────────────
# Routes — Gemini AI Settings
# ─────────────────────────────────────────────────────────────
@api.get("/settings/gemini")
async def get_gemini_settings(admin: dict = Depends(require_admin)):
    cfg = await db.settings.find_one({"key": "gemini"}, {"_id": 0})
    if not cfg:
        return {"api_key": "", "model": "gemini-2.5-flash", "enabled": False}
    cfg.pop("key", None)
    # Mask API key in response (show last 4)
    k = cfg.get("api_key", "")
    if k and len(k) > 8:
        cfg["api_key_masked"] = "***" + k[-4:]
    return cfg

@api.put("/settings/gemini")
async def set_gemini_settings(body: GeminiSettings, admin: dict = Depends(require_admin)):
    doc = body.model_dump()
    doc["key"] = "gemini"
    doc["updated_at"] = now_iso()
    doc["updated_by"] = admin["name"]
    await db.settings.update_one({"key": "gemini"}, {"$set": doc}, upsert=True)
    return {"ok": True}

@api.post("/settings/gemini/test")
async def test_gemini(admin: dict = Depends(require_admin)):
    cfg = await db.settings.find_one({"key": "gemini"}, {"_id": 0})
    if not cfg:
        return {"ok": False, "error": "Save API key first"}
    return await gemini_test_key(cfg.get("api_key", ""))

# ─────────────────────────────────────────────────────────────
# Routes — Telegram Webhook + Invite Link Tracking
# ─────────────────────────────────────────────────────────────
@api.post("/settings/telegram/setup-webhook")
async def setup_webhook(request: Request, admin: dict = Depends(require_admin)):
    cfg = await db.settings.find_one({"key": "telegram"}, {"_id": 0})
    if not cfg or not cfg.get("bot_token"):
        return {"ok": False, "error": "Save bot token first"}
    base = str(request.base_url).rstrip("/")
    if base.startswith("http://") and "localhost" not in base and "127.0.0.1" not in base:
        base = base.replace("http://", "https://", 1)
    webhook_url = f"{base}/api/telegram/webhook"
    res = await tg_set_webhook(cfg["bot_token"], webhook_url)
    if res.get("ok"):
        me = await tg_get_me(cfg["bot_token"])
        bot_user = me.get("result", {}).get("username") if me.get("ok") else None
        await db.settings.update_one(
            {"key": "telegram"},
            {"$set": {"bot_username": bot_user, "webhook_url": webhook_url}},
        )
        return {"ok": True, "webhook_url": webhook_url, "bot_username": bot_user}
    return res

@api.get("/settings/telegram/webhook-info")
async def webhook_info(admin: dict = Depends(require_admin)):
    cfg = await db.settings.find_one({"key": "telegram"}, {"_id": 0})
    if not cfg or not cfg.get("bot_token"):
        return {"ok": False, "error": "Save bot token first"}
    return await tg_get_webhook_info(cfg["bot_token"])

@api.post("/telegram/webhook")
async def telegram_webhook(request: Request):
    """Public — Telegram POSTs updates here. Handles chat_join_request + my_chat_member."""
    try:
        update = await request.json()
    except Exception:
        return {"ok": True}

    log.info(f"Telegram update keys: {list(update.keys())}")
    cfg = await db.settings.find_one({"key": "telegram"}, {"_id": 0})
    token = (cfg or {}).get("bot_token", "") if cfg else ""

    if "my_chat_member" in update:
        mcm = update["my_chat_member"]
        chat = mcm.get("chat", {})
        new_status = mcm.get("new_chat_member", {}).get("status")
        if new_status in ("administrator", "member"):
            await db.channels.update_one(
                {"chat_id": chat.get("id")},
                {"$set": {
                    "chat_id": chat.get("id"),
                    "title": chat.get("title"),
                    "username": chat.get("username"),
                    "type": chat.get("type"),
                    "bot_status": new_status,
                    "added_at": now_iso(),
                }},
                upsert=True,
            )
            log.info(f"Bot is {new_status} in chat: {chat.get('title')} ({chat.get('id')})")
        elif new_status in ("left", "kicked"):
            await db.channels.update_one(
                {"chat_id": chat.get("id")},
                {"$set": {"bot_status": new_status, "removed_at": now_iso()}},
            )

    if "chat_join_request" in update:
        req = update["chat_join_request"]
        chat = req.get("chat", {})
        chat_id = chat.get("id")
        from_user = req.get("from", {})
        user_id = from_user.get("id")
        link_url = (req.get("invite_link") or {}).get("invite_link", "")

        if chat_id and user_id and link_url:
            link = await db.invite_links.find_one({"link_url": link_url})
            if link:
                # Record member entry (idempotent)
                await db.link_members.update_one(
                    {"link_id": link["id"], "telegram_user_id": user_id},
                    {"$set": {
                        "link_id": link["id"],
                        "channel_id": chat_id,
                        "telegram_user_id": user_id,
                        "first_name": from_user.get("first_name", ""),
                        "username": from_user.get("username"),
                        "joined_at": now_iso(),
                        "left_at": None,
                    }},
                    upsert=True,
                )
                await db.invite_links.update_one(
                    {"id": link["id"]},
                    {"$inc": {"members_joined": 1}, "$set": {"last_join_at": now_iso()}},
                )
                if token:
                    await tg_approve_join(token, chat_id, user_id)
                log.info(f"+1 member via link {link.get('link_name')}")

    # Track member leaves
    if "chat_member" in update:
        cm = update["chat_member"]
        chat_id = cm.get("chat", {}).get("id")
        new_status = cm.get("new_chat_member", {}).get("status")
        user_id = cm.get("new_chat_member", {}).get("user", {}).get("id")
        if chat_id and user_id and new_status in ("left", "kicked", "restricted"):
            res = await db.link_members.update_many(
                {"channel_id": chat_id, "telegram_user_id": user_id, "left_at": None},
                {"$set": {"left_at": now_iso()}},
            )
            if res.modified_count:
                log.info(f"Member {user_id} left chat {chat_id} ({res.modified_count} link entries marked)")

    return {"ok": True}

# ─────────────────────────────────────────────────────────────
# Routes — Channels (Admin-managed)
# ─────────────────────────────────────────────────────────────
async def _resolve_chat(token: str, chat_id_or_username: str) -> dict:
    """Try to resolve a chat by ID or @username via Telegram getChat."""
    return await tg_get_chat(token, chat_id_or_username)

def _extract_chat_id_from_invite(invite_link: str) -> Optional[str]:
    """Best-effort: extract chat info hint from invite link.
    t.me/+abc123  → private (need getChat with link)
    t.me/channelname → public, can use @channelname
    """
    if not invite_link:
        return None
    s = invite_link.strip()
    s = s.replace("https://", "").replace("http://", "")
    if s.startswith("t.me/"):
        s = s[5:]
    if s.startswith("+"):
        return None  # private — admin must use chat_id directly
    return "@" + s.split("/")[0].split("?")[0]

@api.get("/channels")
async def list_channels(admin: dict = Depends(require_admin)):
    rows = await db.channels.find({}, {"_id": 0}).sort("added_at", -1).to_list(500)
    return rows

@api.get("/channels/available")
async def available_channels(p: dict = Depends(require_user)):
    """Channels users can pick from (bot must be admin in them)."""
    rows = await db.channels.find(
        {"enabled": {"$ne": False}, "bot_status": {"$in": ["administrator", "member"]}},
        {"_id": 0},
    ).sort("title", 1).to_list(500)
    return rows

@api.post("/channels")
async def add_channel(body: ChannelIn, admin: dict = Depends(require_admin)):
    cfg = await db.settings.find_one({"key": "telegram"}, {"_id": 0})
    if not cfg or not cfg.get("bot_token"):
        raise HTTPException(400, "Configure Telegram bot first")
    token = cfg["bot_token"]

    # Determine lookup key
    lookup = None
    if body.chat_id:
        lookup = body.chat_id.strip()
    elif body.invite_link:
        lookup = _extract_chat_id_from_invite(body.invite_link)
        if not lookup:
            raise HTTPException(400, "Private invite link — please use numeric chat_id instead (bot logs / @userinfobot)")
    else:
        raise HTTPException(400, "Provide chat_id or invite_link")

    res = await tg_get_chat(token, lookup)
    if not res.get("ok"):
        err = res.get("error", "")
        raise HTTPException(400, f"Could not find chat: {err}. Make sure bot is admin in that channel.")

    chat = res["result"]
    chat_id = chat["id"]
    title = body.title_override or chat.get("title") or chat.get("username") or str(chat_id)

    # Check bot status
    try:
        async with httpx.AsyncClient(timeout=15) as cli:
            r = await cli.post(
                f"https://api.telegram.org/bot{token}/getChatMember",
                json={"chat_id": chat_id, "user_id": (await tg_get_me(token)).get("result", {}).get("id")},
            )
            mem_data = r.json()
            bot_status = mem_data.get("result", {}).get("status", "unknown") if mem_data.get("ok") else "unknown"
    except Exception:
        bot_status = "unknown"

    doc = {
        "id": str(uuid.uuid4()),
        "chat_id": chat_id,
        "title": title,
        "original_title": chat.get("title"),
        "username": chat.get("username"),
        "type": chat.get("type"),
        "default_worker_type": body.default_worker_type,
        "notes": body.notes,
        "bot_status": bot_status,
        "enabled": True,
        "added_by": admin["name"],
        "added_at": now_iso(),
    }
    try:
        await db.channels.insert_one(doc)
    except Exception:
        raise HTTPException(409, "Channel already added")
    doc.pop("_id", None)
    return doc

@api.put("/channels/{cid}")
async def edit_channel(cid: str, body: ChannelPatch, admin: dict = Depends(require_admin)):
    ch = await db.channels.find_one({"id": cid}, {"_id": 0})
    if not ch:
        raise HTTPException(404, "Channel not found")
    upd = {k: v for k, v in body.model_dump().items() if v is not None}
    if "title_override" in upd:
        upd["title"] = upd.pop("title_override")
    upd["updated_at"] = now_iso()
    upd["updated_by"] = admin["name"]
    await db.channels.update_one({"id": cid}, {"$set": upd})
    return {"ok": True}

@api.delete("/channels/{cid}")
async def delete_channel(cid: str, admin: dict = Depends(require_admin)):
    ch = await db.channels.find_one({"id": cid}, {"_id": 0})
    if not ch:
        raise HTTPException(404, "Channel not found")
    # Delete related links
    await db.invite_links.delete_many({"channel_id": ch["chat_id"]})
    await db.link_members.delete_many({"channel_id": ch["chat_id"]})
    await db.channels.delete_one({"id": cid})
    return {"ok": True}

# ─────────────────────────────────────────────────────────────
# Routes — Invite Links (use admin's channel list)
# ─────────────────────────────────────────────────────────────
@api.get("/links/me")
async def my_links(p: dict = Depends(require_user)):
    return await db.invite_links.find({"staff_user_id": p["sub"]}, {"_id": 0}).sort("created_at", -1).to_list(100)

@api.get("/links")
async def all_links(admin: dict = Depends(require_admin)):
    return await db.invite_links.find({}, {"_id": 0}).sort("created_at", -1).to_list(2000)

@api.post("/links/create")
async def create_link(body: dict, p: dict = Depends(require_user)):
    """User picks a channel_id (from admin's list)."""
    channel_id = body.get("channel_id")
    name = (body.get("name") or "").strip()
    if not channel_id:
        raise HTTPException(400, "Select a channel")
    ch = await db.channels.find_one({"id": channel_id}, {"_id": 0})
    if not ch:
        raise HTTPException(404, "Channel not found in admin list")
    if ch.get("enabled") is False:
        raise HTTPException(400, "Channel is disabled by admin")

    cfg = await db.settings.find_one({"key": "telegram"}, {"_id": 0})
    if not cfg or not cfg.get("bot_token"):
        raise HTTPException(400, "Telegram bot not configured")
    token = cfg["bot_token"]
    user = await db.users.find_one({"id": p["sub"]}, {"_id": 0})
    link_name = (name or f"{user['name']}'s Link")[:32]
    res = await tg_create_invite(token, ch["chat_id"], name=link_name, creates_join_request=True)
    if not res.get("ok"):
        err = res.get("error", "")
        if "rights" in err.lower() or "not enough" in err.lower():
            raise HTTPException(400, "Bot lacks 'Invite Users' permission in this channel")
        raise HTTPException(400, f"Could not create link: {err}")

    invite = res["result"]
    doc = {
        "id": str(uuid.uuid4()),
        "staff_user_id": p["sub"],
        "staff_username": user.get("username"),
        "staff_name": user["name"],
        "staff_telegram": user["telegram"],
        "channel_id": ch["chat_id"],
        "channel_db_id": ch["id"],
        "channel_title": ch["title"],
        "link_url": invite["invite_link"],
        "link_name": link_name,
        "members_joined": 0,
        "created_at": now_iso(),
        "active": True,
    }
    await db.invite_links.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api.delete("/links/{lid}")
async def delete_link(lid: str, p: dict = Depends(require_user)):
    link = await db.invite_links.find_one({"id": lid, "staff_user_id": p["sub"]}, {"_id": 0})
    if not link:
        raise HTTPException(404, "Link not found or not yours")
    cfg = await db.settings.find_one({"key": "telegram"}, {"_id": 0})
    if cfg and cfg.get("bot_token"):
        await tg_revoke_invite(cfg["bot_token"], link["channel_id"], link["link_url"])
    await db.invite_links.delete_many({"id": lid})
    await db.link_members.delete_many({"link_id": lid})
    return {"ok": True}

@api.get("/links/{lid}/members")
async def link_members(lid: str, p: dict = Depends(get_principal)):
    """Live members of a link with join/leave status."""
    link = await db.invite_links.find_one({"id": lid}, {"_id": 0})
    if not link:
        raise HTTPException(404, "Not found")
    if p["role"] == "user" and link["staff_user_id"] != p["sub"]:
        raise HTTPException(403, "Not yours")
    rows = await db.link_members.find({"link_id": lid}, {"_id": 0}).sort("joined_at", -1).to_list(2000)
    active = sum(1 for r in rows if r.get("left_at") is None)
    left = len(rows) - active
    return {"link": link, "active": active, "left": left, "members": rows}

# ─────────────────────────────────────────────────────────────
# Daily summary scheduler (12 PM IST)
# ─────────────────────────────────────────────────────────────
IST = pytz.timezone("Asia/Kolkata")
scheduler: Optional[AsyncIOScheduler] = None

async def run_daily_summary_job(force: bool = False):
    cfg = await tg_get_config(db)
    if not cfg or not cfg.get("enabled"):
        log.info("Daily summary skipped: telegram disabled")
        return
    if not force and not cfg.get("daily_summary", True):
        log.info("Daily summary skipped: daily_summary off")
        return
    now_utc = datetime.now(timezone.utc)
    since_utc = now_utc - timedelta(hours=24)
    rows = await db.reports.find(
        {"created_at": {"$gte": since_utc.isoformat()}},
        {"_id": 0},
    ).sort("created_at", -1).to_list(2000)
    date_str = datetime.now(IST).strftime("%d %b %Y")
    msg = build_daily_summary(rows, date_str)
    res = await tg_send(db, msg)
    log.info(f"Daily summary sent: {res}")


async def run_daily_auto_reports(force: bool = False):
    """At 10 AM IST: for each user, count yesterday's NET active members
    (joined yesterday AND still in chat at 10 AM today) per link, calculate
    salary using their default_worker_type, create AUTO report.
    """
    log.info("Running 10 AM IST auto-report job...")
    now_ist = datetime.now(IST)
    period_end = now_ist.replace(hour=10, minute=0, second=0, microsecond=0)
    if now_ist < period_end:
        period_end = period_end - timedelta(days=1)
    period_start = period_end - timedelta(days=1)
    period_end_utc = period_end.astimezone(timezone.utc).isoformat()
    period_start_utc = period_start.astimezone(timezone.utc).isoformat()
    log.info(f"Period: {period_start_utc} → {period_end_utc}")

    users = await db.users.find({"username": {"$exists": True}}, {"_id": 0}).to_list(1000)
    reports_created = 0
    for user in users:
        # All this user's links
        links = await db.invite_links.find({"staff_user_id": user["id"]}, {"_id": 0}).to_list(100)
        if not links:
            continue
        link_ids = [l["id"] for l in links]
        # Net active = joined in period AND (left_at is null OR left_at > period_end)
        active_count = await db.link_members.count_documents({
            "link_id": {"$in": link_ids},
            "joined_at": {"$gte": period_start_utc, "$lt": period_end_utc},
            "$or": [{"left_at": None}, {"left_at": {"$gte": period_end_utc}}],
        })
        total_joined = await db.link_members.count_documents({
            "link_id": {"$in": link_ids},
            "joined_at": {"$gte": period_start_utc, "$lt": period_end_utc},
        })
        left_count = total_joined - active_count
        if total_joined == 0:
            continue  # skip users with no joins in period
        worker_type = user.get("default_worker_type", "5 ID Worker")
        salary, slab = await calc_salary(worker_type, active_count)
        doc = {
            "id": str(uuid.uuid4()),
            "user_id": user["id"],
            "name": user["name"],
            "telegram": user.get("telegram", ""),
            "worker_type": worker_type,
            "member_count": active_count,
            "ai_count": active_count,
            "joined_count": total_joined,
            "left_count": left_count,
            "source": "AUTO_10AM",
            "channel_title": "Multiple" if len(links) > 1 else links[0]["channel_title"],
            "salary": salary,
            "status": "VERIFIED",
            "slab_label": slab,
            "period_start": period_start_utc,
            "period_end": period_end_utc,
            "created_at": now_iso(),
        }
        await db.reports.insert_one(doc)
        reports_created += 1
        log.info(f"AUTO report for {user['name']}: {active_count} active ({left_count} left), ₹{salary}")
        # Telegram notify
        try:
            tg_cfg = await tg_get_config(db)
            if tg_cfg and tg_cfg.get("enabled") and tg_cfg.get("notify_on_report", True):
                doc.pop("_id", None)
                await tg_send(db, build_report_notification(doc))
        except Exception:
            pass
    log.info(f"Auto-report job complete: {reports_created} reports created")


@api.post("/settings/run-auto-reports-now")
async def trigger_auto_reports(admin: dict = Depends(require_admin)):
    """Manual trigger for the 10 AM auto-report job."""
    await run_daily_auto_reports(force=True)
    return {"ok": True}

# ─────────────────────────────────────────────────────────────
# App wiring
# ─────────────────────────────────────────────────────────────
app.include_router(api)

app.add_middleware(
    CORSMiddleware,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("startup")
async def on_start():
    await ensure_indexes()
    await seed_admin()
    await seed_slabs()
    global scheduler
    scheduler = AsyncIOScheduler(timezone=IST)
    scheduler.add_job(
        run_daily_auto_reports,
        CronTrigger(hour=10, minute=0, timezone=IST),
        id="daily_auto_reports",
        replace_existing=True,
    )
    scheduler.add_job(
        run_daily_summary_job,
        CronTrigger(hour=12, minute=0, timezone=IST),
        id="daily_salary_summary",
        replace_existing=True,
    )
    scheduler.start()
    log.info("TeamCrazy Hub API ready · Auto-reports @ 10 AM IST · Telegram summary @ 12 PM IST")

@app.on_event("shutdown")
async def on_stop():
    if scheduler:
        scheduler.shutdown(wait=False)
    client.close()
